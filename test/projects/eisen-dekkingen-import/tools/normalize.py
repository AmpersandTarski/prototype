#!/usr/bin/env python3
"""Dunne, TIJDELIJKE normalizer: ont-kruist de platte join ESE ⋈ Organisme ⋈ Kenmerk ⋈ Dekking
terug tot de genormaliseerde tabellen en schrijft ze in het Ampersand-INTERFACE-xlsx-formaat
(sheet-titel == interface-label, kolomkop == leaf-label, kolom A == atoom-id). Beide importers
lezen dat formaat identiek (compile-time `INCLUDE`, runtime `/admin/import`).

De onregelmatigheden die deze normalizer gladstrijkt (cartesiaanse duplicatie, ontbrekende
sleutels voor EKK/DKG, dubbele DEKKING_CODE-kolom, ';'-meervoud) zijn export-artefacten van de
Oracle-query — geen modelkenmerken. Als de bron ooit genormaliseerd wordt aangeleverd, vervalt
deze normalizer.

  python3 tools/normalize.py <sample-raw.json> <out-dir> [--split]
Zonder --split: één workbook out-dir/populatie.xlsx met alle ESE's.
Met  --split: out-dir/compile.xlsx (even index) + out-dir/runtime.xlsx (oneven index).
"""
import json, sys, os
from collections import defaultdict, OrderedDict
import openpyxl

SEP = "␟"  # unit separator voor samengestelde Product-sleutel (komt niet in data voor)

def build(header, rows):
    ix = {h: i for i, h in enumerate(header)}
    def g(r, name):
        v = r[ix[name]]
        return v if (v is not None and not (isinstance(v, str) and v.strip() == "")) else None

    ese = OrderedDict()      # ese_id -> dict
    product = {}             # pid -> (naam, groep)
    land = {}                # code -> naam
    organisme = {}           # oge -> naam
    verkl_map = {}           # (concept,taal,tekst) -> vk_id
    verkl = OrderedDict()    # vk_id -> (concept,taal,tekst)
    kenmerk = OrderedDict()  # k_id -> dict   (per-ese distinct EKK-tupel)
    dekking = OrderedDict()  # d_id -> dict   (per-ese distinct DKG-scalartupel; verklaring multi)
    ese_ekk_seen = defaultdict(dict)   # ese -> ekk-tuple -> k_id
    ese_dkg_seen = defaultdict(dict)   # ese -> dkg-scalartuple -> d_id

    def vk_for(concept, taal, tekst):
        if concept is None and taal is None and tekst is None:
            return None
        key = (concept, taal, tekst)
        if key not in verkl_map:
            vid = f"VK{len(verkl_map) + 1}"
            verkl_map[key] = vid
            verkl[vid] = key
        return verkl_map[key]

    for r in rows:
        eid = str(g(r, "ESE_ESE_ID"))
        if eid == "None":
            continue
        if eid not in ese:
            groep = g(r, "PRODUCT_GROEP"); pnaam = g(r, "PRODUCT_NAAM")
            pid = None
            if pnaam is not None:
                pid = f"{groep}{SEP}{pnaam}" if groep is not None else pnaam
                product[pid] = (pnaam, groep)
            code = g(r, "LAND_CODE")
            if code is not None:
                land[str(code)] = g(r, "LAND_NAAM")
            ese[eid] = {
                "product": pid,
                "land": None if code is None else str(code),
                "bak": None if g(r, "BAK_ID") is None else str(g(r, "BAK_ID")),
                "status": g(r, "ESE_STATUS"),
                "hoofdeis": None if g(r, "HOOFDEIS") is None else str(g(r, "HOOFDEIS")),
                "eisNaam": g(r, "ESE_EIS_NAAM"),
                "typeEis": g(r, "SES_TYPE_EIS_NAAM"),
                "bron": g(r, "ESE_BRON"),
                "interneMemo": g(r, "ESE_INTERNE_MEMO"),
                "externeMemo": g(r, "ESE_EXTERNE_MEMO"),
                "beginDatum": g(r, "ESE_BEGIN_DATUM"),
                "eindDatum": g(r, "ESE_EIND_DATUM"),
                "normOperator": g(r, "ESE_NORM_OPERATOR"),
                "normWaarde": g(r, "ESE_NORM_WAARDE"),
                "normEenheid": g(r, "ESE_NORM_EENHEID"),
                "dekkingCode": g(r, "DEKKING_CODE"),   # ruwe string, ';'-meervoud
                "organisme": set(),
                "verklaring": set(),
            }
        e = ese[eid]
        # organisme (meervoudig)
        oge = g(r, "OGE_ID")
        if oge is not None:
            organisme[str(oge)] = g(r, "ESE_ORGANISME")
            e["organisme"].add(str(oge))
        # ESE-verklaring (meervoudig)
        vk = vk_for(g(r, "ESE_VERKLARINGSCONCEPT"), g(r, "ESE_TAAL"), g(r, "ESE_VERKLARINGSTEKST"))
        if vk is not None:
            e["verklaring"].add(vk)
        # kenmerk (EKK): distinct volledige tupel per ESE
        ekk = (g(r, "EKK_NAAM"), g(r, "EKK_WAARDE_ORG"), g(r, "EKK_WAARDE_VERTAALD"),
               g(r, "EKK_STATUS"), g(r, "EKK_EIND_DATUM"))
        if any(x is not None for x in ekk):
            seen = ese_ekk_seen[eid]
            if ekk not in seen:
                kid = f"K{eid}_{len(seen) + 1}"
                seen[ekk] = kid
                kenmerk[kid] = {"kenmerkVanEse": eid, "kenmerksoort": ekk[0],
                                "waardeOrigineel": ekk[1], "waardeVertaald": ekk[2],
                                "status": ekk[3], "eindDatum": ekk[4]}
        # dekking (DKG): distinct scalartupel per ESE; verklaring apart (meervoudig)
        dkey = (g(r, "SETNR"), g(r, "ALT"), g(r, "VOLGNR"), g(r, "DKG_STATUS"),
                g(r, "DKG_EIND_DATUM"), g(r, "DKG_TYPE_DEKKING"), g(r, "DKG_CODE_INSPECTIE"),
                g(r, "DKG_NAAM_DEKKING"), g(r, "DKG_MODULE_INSPECTIE"),
                g(r, "DKG_NORM_OPERATOR"), g(r, "DKG_NORM_WAARDE"), g(r, "DKG_NORM_EENHEID"))
        if any(x is not None for x in dkey):
            seen = ese_dkg_seen[eid]
            if dkey not in seen:
                did = f"D{eid}_{len(seen) + 1}"
                seen[dkey] = did
                dekking[did] = {
                    "dekkingVanEse": eid,
                    "setnummer": dkey[0], "alternatief": dkey[1], "volgnummer": dkey[2],
                    "status": dkey[3], "eindDatum": dkey[4], "typeDekking": dkey[5],
                    "codeInspectie": dkey[6], "naamDekking": dkey[7], "moduleInspectie": dkey[8],
                    "normOperator": dkey[9], "normWaarde": dkey[10], "normEenheid": dkey[11],
                    "verklaring": set(),
                }
            did = seen[dkey]
            vkd = vk_for(g(r, "DKG_VERKLARINGSCONCEPT"), g(r, "DKG_VER_CONCEPT_TAAL"),
                         g(r, "DKG_VERKLARINGSTEKST"))
            if vkd is not None:
                dekking[did]["verklaring"].add(vkd)

    return dict(ese=ese, product=product, land=land, organisme=organisme, verkl=verkl,
                kenmerk=kenmerk, dekking=dekking)

# sheet-definities: (interface/concept, [(kolomkop, veld, mode)]) mode: 'v' scalar, 'm' multi(','),
# 'M;' multi(';'). kolom A wordt apart geschreven.
def cell(v):
    return "" if v is None else str(v)

def write_workbook(path, model, ese_ids):
    ese_ids = set(ese_ids)
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    # welke lookup-atomen referenceert deze partitie?
    ref_prod, ref_land, ref_org, ref_vk = set(), set(), set(), set()
    for eid in ese_ids:
        e = model["ese"][eid]
        if e["product"]: ref_prod.add(e["product"])
        if e["land"]: ref_land.add(e["land"])
        ref_org |= e["organisme"]; ref_vk |= e["verklaring"]
    kenmerk = {k: v for k, v in model["kenmerk"].items() if v["kenmerkVanEse"] in ese_ids}
    dekking = {k: v for k, v in model["dekking"].items() if v["dekkingVanEse"] in ese_ids}
    for d in dekking.values():
        ref_vk |= d["verklaring"]

    def sheet(title, concept, colspecs, idrows):
        ws = wb.create_sheet(title)
        ws.cell(1, 1, concept)
        for j, (hdr, _f, _m) in enumerate(colspecs, start=2):
            ws.cell(1, j, hdr)
        ri = 2
        for aid, rec in idrows:
            ws.cell(ri, 1, cell(aid))
            for j, (_h, f, m) in enumerate(colspecs, start=2):
                v = rec.get(f)
                if m == "m":
                    v = ",".join(sorted(v)) if v else ""
                ws.cell(ri, j, cell(v))
            ri += 1

    # ESE
    sheet("ESE", "ESE", [
        ("product", "product", "v"), ("land", "land", "v"), ("bak", "bak", "v"),
        ("status", "status", "v"), ("hoofdeis", "hoofdeis", "v"), ("eisNaam", "eisNaam", "v"),
        ("[organisme,]", "organisme", "m"), ("typeEis", "typeEis", "v"), ("bron", "bron", "v"),
        ("interneMemo", "interneMemo", "v"), ("externeMemo", "externeMemo", "v"),
        ("beginDatum", "beginDatum", "v"), ("eindDatum", "eindDatum", "v"),
        ("[verklaring,]", "verklaring", "m"),
        ("normOperator", "normOperator", "v"), ("normWaarde", "normWaarde", "v"),
        ("normEenheid", "normEenheid", "v"), ("[dekkingCode;]", "dekkingCode", "v"),
    ], [(eid, model["ese"][eid]) for eid in model["ese"] if eid in ese_ids])

    sheet("Product", "Product", [("productnaam", "naam", "v"), ("productgroep", "groep", "v")],
          [(pid, {"naam": model["product"][pid][0], "groep": model["product"][pid][1]})
           for pid in model["product"] if pid in ref_prod])

    sheet("Land", "Land", [("landnaam", "naam", "v")],
          [(c, {"naam": model["land"][c]}) for c in model["land"] if c in ref_land])

    sheet("Organisme", "Organisme", [("organismenaam", "naam", "v")],
          [(o, {"naam": model["organisme"][o]}) for o in model["organisme"] if o in ref_org])

    sheet("Verklaring", "Verklaring", [
        ("verklaringconcept", "concept", "v"), ("verklaringtaal", "taal", "v"),
        ("verklaringtekst", "tekst", "v")],
        [(vid, {"concept": model["verkl"][vid][0], "taal": model["verkl"][vid][1],
                "tekst": model["verkl"][vid][2]}) for vid in model["verkl"] if vid in ref_vk])

    sheet("Kenmerk", "Kenmerk", [
        ("kenmerkVanEse", "kenmerkVanEse", "v"), ("kenmerksoort", "kenmerksoort", "v"),
        ("waardeOrigineel", "waardeOrigineel", "v"), ("waardeVertaald", "waardeVertaald", "v"),
        ("status", "status", "v"), ("eindDatum", "eindDatum", "v")],
        list(kenmerk.items()))

    sheet("Dekking", "Dekking", [
        ("dekkingVanEse", "dekkingVanEse", "v"), ("setnummer", "setnummer", "v"),
        ("alternatief", "alternatief", "v"), ("volgnummer", "volgnummer", "v"),
        ("status", "status", "v"), ("eindDatum", "eindDatum", "v"),
        ("typeDekking", "typeDekking", "v"), ("codeInspectie", "codeInspectie", "v"),
        ("naamDekking", "naamDekking", "v"), ("moduleInspectie", "moduleInspectie", "v"),
        ("[verklaring,]", "verklaring", "m"), ("normOperator", "normOperator", "v"),
        ("normWaarde", "normWaarde", "v"), ("normEenheid", "normEenheid", "v")],
        list(dekking.items()))

    wb.save(path)
    return dict(ese=len(ese_ids), product=len(ref_prod), land=len(ref_land),
                organisme=len(ref_org), verklaring=len(ref_vk),
                kenmerk=len(kenmerk), dekking=len(dekking))

def main():
    src, outdir = sys.argv[1], sys.argv[2]
    split = "--split" in sys.argv[3:]
    data = json.load(open(src))
    model = build(data["header"], data["rows"])
    all_ese = list(model["ese"].keys())
    os.makedirs(outdir, exist_ok=True)
    if split:
        compile_ese = [e for i, e in enumerate(all_ese) if i % 2 == 0]
        runtime_ese = [e for i, e in enumerate(all_ese) if i % 2 == 1]
        s1 = write_workbook(os.path.join(outdir, "compile.xlsx"), model, compile_ese)
        s2 = write_workbook(os.path.join(outdir, "runtime.xlsx"), model, runtime_ese)
        # manifest: welke ESE's zitten in welke partitie (zodat de verify elke route apart toetst).
        # Alleen de ESE-id-splitsing — niet de paar-afleiding — dus de mapping-onafhankelijkheid blijft.
        with open(os.path.join(outdir, "partition.json"), "w") as f:
            json.dump({"compile": compile_ese, "runtime": runtime_ese,
                       "all": all_ese}, f, ensure_ascii=False)
        print("compile.xlsx", s1)
        print("runtime.xlsx", s2)
    else:
        s = write_workbook(os.path.join(outdir, "populatie.xlsx"), model, all_ese)
        print("populatie.xlsx", s)

if __name__ == "__main__":
    main()
